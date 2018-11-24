import openpyxl
import csv
from datetime import timedelta

wb = openpyxl.load_workbook('IHearVoicesData.xlsx', data_only=True)

ws = wb.active

data = []

for row in range(3, 57):
    status = ws.cell(row=row, column=5).value

    if status == 'NO DATA':
        continue
    elif status == 'SAME STATUS':
        data.append({'status': status})
        continue

    time_delta = timedelta(hours=3)

    start = int((ws.cell(row=row, column=6).value + time_delta).timestamp())
    end = int((ws.cell(row=row, column=7).value + time_delta).timestamp())

    data.append({'status': status,
                 'start': start,
                 'end': end})

for i in range(0, len(data)):
    if data[i]['status'] == 'SAME STATUS':
        data[i - 1]['end'] = data[i + 1]['end']
        data[i]['status'] = 'DELETE'
        data[i + 1]['status'] = 'DELETE'

data = [item for item in data if item['status'] != 'DELETE']

with open('statuses.csv', mode='w+', encoding='utf-8', newline='') as csv_file:
    writer = csv.DictWriter(csv_file, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
